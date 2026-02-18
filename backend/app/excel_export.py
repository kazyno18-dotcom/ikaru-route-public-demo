import io
from datetime import date, datetime, timedelta
from sqlalchemy.orm import Session
from sqlalchemy import and_
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter
from app import models

# サービス種別カラーコード（Excel用）
SERVICE_COLORS = {
    "身体": "BDD7EE",   # ライトブルー
    "家事": "C6EFCE",   # ライトグリーン
    "生活": "E2CEFF",   # ライトパープル
    "重度": "FFCCCC",   # ライトレッド
    "障がい": "FFE0B2", # ライトオレンジ
}

TWO_STAFF_COLOR = "FFF2CC"  # 黄色（2人体制）
HEADER_COLOR = "2F4F8F"     # ダークブルー（ヘッダー）


def generate_route_excel(db: Session, target_date: date, include_revenue: bool = False) -> io.BytesIO:
    """現行フォーマット準拠の日次ルート表Excelを生成"""
    wb = Workbook()
    ws = wb.active
    ws.title = f"ルート表_{target_date.strftime('%m月%d日')}"

    # 印刷設定（A3横向き）
    ws.page_setup.paperSize = 8  # A3
    ws.page_setup.orientation = "landscape"
    ws.oddHeader.center.text = f"鶴進HMG 訪問介護ルート表　{target_date.strftime('%Y年%m月%d日')}"
    ws.oddFooter.right.text = "ページ &P / &N"

    # スタッフ一覧取得
    staff_list = db.query(models.Staff).filter(models.Staff.is_active == True).order_by(models.Staff.name).all()

    # 時間軸設定（5:00〜翌5:00、15分刻み）
    start_hour = 5
    time_slots = []
    for h in range(24):
        for m in [0, 15, 30, 45]:
            actual_hour = (start_hour + h) % 24
            time_slots.append(f"{actual_hour:02d}:{m:02d}")

    # ヘッダー行
    ws.cell(row=1, column=1, value="時刻")
    ws.cell(row=1, column=1).fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
    ws.cell(row=1, column=1).font = Font(color="FFFFFF", bold=True)
    ws.column_dimensions["A"].width = 8

    for col_idx, staff in enumerate(staff_list, start=2):
        cell = ws.cell(row=1, column=col_idx, value=staff.name)
        cell.fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    # 時間軸列
    for row_idx, time_str in enumerate(time_slots, start=2):
        cell = ws.cell(row=row_idx, column=1, value=time_str)
        cell.alignment = Alignment(horizontal="center")
        if time_str.endswith(":00"):
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # 訪問データを取得してセルに書き込み
    visits = db.query(models.Visit).filter(
        models.Visit.date == target_date
    ).all()

    # スタッフ別訪問マップ
    staff_visits = {str(s.staff_id): [] for s in staff_list}
    for visit in visits:
        if visit.staff_id and str(visit.staff_id) in staff_visits:
            staff_visits[str(visit.staff_id)].append(visit)

    for col_idx, staff in enumerate(staff_list, start=2):
        staff_visit_list = staff_visits.get(str(staff.staff_id), [])
        total_minutes = 0
        visit_count = 0

        for visit in staff_visit_list:
            # 開始時刻の行を特定
            visit_start = visit.scheduled_start
            start_slot_hour = (visit_start.hour - start_hour) % 24
            start_slot_min = visit_start.minute // 15
            row_idx = 2 + start_slot_hour * 4 + start_slot_min

            # セル内容
            duration = int((visit.scheduled_end - visit.scheduled_start).total_seconds() / 60)
            client_name = visit.client.name if visit.client else "不明"
            service_type = visit.service_type if isinstance(visit.service_type, str) else visit.service_type.value
            content = f"{client_name} {service_type}{duration}"

            if visit.companion_staff_id and visit.visit_type == "two_staff":
                companion = db.query(models.Staff).filter(
                    models.Staff.staff_id == visit.companion_staff_id
                ).first()
                if companion:
                    content += f"（{companion.name}と）"

            cell = ws.cell(row=row_idx, column=col_idx, value=content)

            # セル色分け
            color = SERVICE_COLORS.get(service_type, "FFFFFF")
            if visit.visit_type == "two_staff":
                color = TWO_STAFF_COLOR
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell.alignment = Alignment(wrap_text=True, vertical="top")

            total_minutes += duration
            visit_count += 1

        # 集計行（最終行）
        summary_row = 2 + len(time_slots)
        ws.cell(row=summary_row, column=1, value="集計")
        summary_cell = ws.cell(
            row=summary_row,
            column=col_idx,
            value=f"{total_minutes // 60}時間{total_minutes % 60}分 / {visit_count}件"
        )
        summary_cell.font = Font(bold=True)
        summary_cell.fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")

    # 売上集計シート（権限付き）
    if include_revenue:
        ws_rev = wb.create_sheet(title="売上集計")
        _generate_revenue_sheet(ws_rev, db, target_date, staff_list)

    # 進捗率シート
    ws_prog = wb.create_sheet(title="進捗率")
    _generate_progress_sheet(ws_prog, db, target_date, staff_list)

    # オートフィルタ設定
    ws.auto_filter.ref = f"A1:{get_column_letter(len(staff_list) + 1)}1"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _generate_revenue_sheet(ws, db: Session, target_date: date, staff_list):
    """売上集計シート生成"""
    ws.cell(row=1, column=1, value="スタッフ名").font = Font(bold=True)
    ws.cell(row=1, column=2, value="訪問件数").font = Font(bold=True)
    ws.cell(row=1, column=3, value="売上合計（円）").font = Font(bold=True)
    ws.cell(row=1, column=4, value="稼働時間").font = Font(bold=True)

    for row_idx, staff in enumerate(staff_list, start=2):
        revenues = db.query(models.Revenue).filter(
            and_(
                models.Revenue.staff_id == staff.staff_id,
                models.Revenue.date == target_date
            )
        ).all()
        total_revenue = sum(r.amount for r in revenues)
        total_minutes = sum(r.duration_minutes for r in revenues)

        ws.cell(row=row_idx, column=1, value=staff.name)
        ws.cell(row=row_idx, column=2, value=len(revenues))
        ws.cell(row=row_idx, column=3, value=total_revenue)
        ws.cell(row=row_idx, column=4, value=f"{total_minutes // 60}時間{total_minutes % 60}分")


def _generate_progress_sheet(ws, db: Session, target_date: date, staff_list):
    """進捗率シート生成"""
    ws.cell(row=1, column=1, value="スタッフ名").font = Font(bold=True)
    ws.cell(row=1, column=2, value="担当件数").font = Font(bold=True)
    ws.cell(row=1, column=3, value="完了件数").font = Font(bold=True)
    ws.cell(row=1, column=4, value="中止件数").font = Font(bold=True)
    ws.cell(row=1, column=5, value="未実施件数").font = Font(bold=True)
    ws.cell(row=1, column=6, value="完了率").font = Font(bold=True)

    for row_idx, staff in enumerate(staff_list, start=2):
        visits = db.query(models.Visit).filter(
            and_(
                models.Visit.staff_id == staff.staff_id,
                models.Visit.date == target_date
            )
        ).all()
        total = len(visits)
        completed = sum(1 for v in visits if v.status == models.VisitStatusEnum.completed)
        cancelled = sum(1 for v in visits if v.status == models.VisitStatusEnum.cancelled)
        not_done = sum(1 for v in visits if v.status == models.VisitStatusEnum.not_done)
        rate = round(completed / total * 100, 1) if total > 0 else 0

        ws.cell(row=row_idx, column=1, value=staff.name)
        ws.cell(row=row_idx, column=2, value=total)
        ws.cell(row=row_idx, column=3, value=completed)
        ws.cell(row=row_idx, column=4, value=cancelled)
        ws.cell(row=row_idx, column=5, value=not_done)
        ws.cell(row=row_idx, column=6, value=f"{rate}%")
